import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

#Ajustando a aparência padrão do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema de Cadastro de Fornecedores")
        self.geometry("700x500")
    
    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', '#fff']).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=10, y=460)
        
        
    def todo_sistema(self):
            frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
            frame.place(x=0, y=10)
            title = ctk.CTkLabel(frame, text="Sistema de Cadastro de fornecedores", font=("Century Gothic bold", 24), text_color="#000").place(x=160, y=15)
            span = ctk.CTkLabel(self, text="Preencha todos os campos do formulário ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)
            
            ficheiro = pathlib.Path("Entradas.xlsx")
                
            if ficheiro.exists():
                pass
            else:
                ficheiro=Workbook()
                folha = ficheiro.active
                folha['A1'] = "Nome Completo"
                folha['B1'] = "Fornecedor"
                folha['C1'] = "Placa"
                folha['D1'] = "Local de descarga"
                folha['E1'] = "Nota fiscal"
                folha['F1'] = "Observações"
                    
                ficheiro.save("Entradas.xlsx")
                
            def submit():
                
               
                
                #pegando os dados dos entrys
                name = name_value.get()
                placa = placa_value.get()
                nf = nf_value.get()
                local = local_combobox.get()
                fornecedor = fornecedor_combobox.get()
                obs = obs_entry.get(0.0, END)
                
                if (name =="" or fornecedor=="" or placa =="" or nf==""):
                   messagebox.showerror("Sistema", "Erro\nPor favor preencha todos os campos.")
                else:
                
                    ficheiro = openpyxl.load_workbook('Entradas.xlsx')
                    folha = ficheiro.active
                    folha.cell(column=1, row=folha.max_row+1,value=name)
                    folha.cell(column=2, row=folha.max_row,value=fornecedor)
                    folha.cell(column=3, row=folha.max_row,value=placa)
                    folha.cell(column=4, row=folha.max_row,value=local)
                    folha.cell(column=5, row=folha.max_row,value=nf)
                    folha.cell(column=6, row=folha.max_row,value=obs)
                
                    ficheiro.save(r"Entradas.xlsx")
                    messagebox.showinfo("Sistema", "Dados salvos com sucesso!")

            def clear():
                name = name_value.set("")
                placa = placa_value.set("")
                nf = nf_value.set("")
                obs = obs_entry.delete(0.0, END)
               
            #Texts variables
            name_value = StringVar()
            placa_value = StringVar()
            nf_value = StringVar()
            
            #Entrys
            name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 16), fg_color="transparent")
            placa_entry = ctk.CTkEntry(self, width=150, textvariable=placa_value, font=("Century Gothic", 16), fg_color="transparent")
            nf_entry = ctk.CTkEntry(self, width=200, textvariable=nf_value, font=("Century Gothic", 16), fg_color="transparent")
            
            #Combobox
            local_combobox = ctk.CTkComboBox(self, values=["Produção", "Logística"], font=("Century Gothic bold", 14))
            fornecedor_combobox = ctk.CTkComboBox(self, values=["Fornecedor 1", "Fornecedor 2", "Fornecedor 3", "Fornecedor 4"], font=("Century Gothic bold", 14))
            local_combobox.set("Produção")
            
            #Entrada de observações
            obs_entry = ctk.CTkTextbox(self, width=450, height=60, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")
            
            #Labels
            lb_name = ctk.CTkLabel(self, text="Nome completo: ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
            lb_placa = ctk.CTkLabel(self, text="Placa da carreta: ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
            lb_local = ctk.CTkLabel(self, text="Local ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
            lb_fornecedor = ctk.CTkLabel(self, text="Fornecedor ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
            lb_nf = ctk.CTkLabel(self, text="Nota fiscal ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
            lb_obs = ctk.CTkLabel(self, text="Observação ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
            
            btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420)
            btn_submit = ctk.CTkButton(self, text="Limpar campos".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=420)
            
            #Posicionamento os elementos na janela
            lb_name.place(x=50, y=110)
            name_entry.place(x=50, y=140)
            
            lb_placa.place(x=450, y=110)
            placa_entry.place(x=450, y=140)
            
            lb_local.place(x=250, y=190)
            local_combobox.place(x=250, y=220)
            
            lb_fornecedor.place(x=50, y=190)
            fornecedor_combobox.place(x=50, y=220)
            
            lb_nf.place(x=450, y=190)
            nf_entry.place(x=450, y=220)
            
            lb_obs.place(x=50, y=250)
            obs_entry.place(x=50, y=280)
            
        
    
    
    
    def change_apm(self, new_apperance_mode):
        ctk.set_appearance_mode(new_apperance_mode)


if __name__=="__main__":
    app = App()
    app.mainloop()