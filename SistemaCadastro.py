from tkinter import *
from tkinter import messagebox
import customtkinter as ctk
import openpyxl
import pathlib
from openpyxl import Workbook, load_workbook
from tkinter import ttk
from datetime import datetime
import subprocess

# Ajustando a aparência padrão do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()
        self.todo_sistema()

    def layout_config(self):
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        window_width = 1280
        window_height = 760
        x_position = (screen_width - window_width) // 2
        y_position = (screen_height - window_height) // 2
        self.title("Sistema de Cadastro de Fornecedores")
        self.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")  # Define a geometria da janela
        
    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', '#fff'])
        self.lb_apm.place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_appearance)
        self.opt_apm.place(x=10, y=460)
        
    def load_data(self):
        # Carregar dados da planilha para a Treeview
        self.carregarDadosTreeView()
        
    def carregarDadosTreeView(self):
        try:
            planilha = openpyxl.load_workbook('Entradas.xlsx')
            folha = planilha.active
            for row in folha.iter_rows(min_row=2, values_only=True):
                self.tree.insert("", "end", values=row)
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo de dados não encontrado.")
        
    def load_fornecedores(self):
        try:
            workbook = openpyxl.load_workbook('Fornecedores.xlsx')
            sheet = workbook.active
            fornecedores = [cell.value for cell in sheet['A'] if cell.value]  # Obtém os valores da coluna A
            return fornecedores
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo de fornecedores não encontrado.")

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=1200, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Cadastro de fornecedores", font=("Century Gothic bold", 24), text_color="#000").place(x=160, y=15)
        span = ctk.CTkLabel(self, text="Preencha todos os campos do formulário ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)
        
        # Adicionando a Treeview
        self.tree = ttk.Treeview(self)
        self.tree.place(x=30, y=520, width=1100, height=150)
        
        # Definindo as colunas da Treeview
        self.tree["columns"] = ("Name", "Fornecedor", "Placa", "Local", "Nota Fiscal", "Hora Chegada", "Hora Inicial", "Hora Final", "Diferença", "Observações", "Peso Inicial", "Tara", "Peso Final")
        
        self.tree.column("#0", width=0, stretch=NO)  # Oculta a primeira coluna
        self.tree.column("Name", anchor=W, width=80)
        self.tree.column("Fornecedor", anchor=W, width=80)
        self.tree.column("Placa", anchor=W, width=100)
        self.tree.column("Local", anchor=W, width=80)
        self.tree.column("Nota Fiscal", anchor=W, width=80)
        self.tree.column("Hora Chegada", anchor=W, width=50)
        self.tree.column("Hora Inicial", anchor=W, width=50)
        self.tree.column("Hora Final", anchor=W, width=50)
        self.tree.column("Diferença", anchor=W, width=60)
        self.tree.column("Observações", anchor=W, width=100)
        self.tree.column("Peso Inicial", anchor=W, width=100)
        self.tree.column("Tara", anchor=W, width=100)
        self.tree.column("Peso Final", anchor=W, width=100)
        
        # Definindo os cabeçalhos das colunas
        self.tree.heading("#0", text="", anchor=W)
        self.tree.heading("Name", text="Nome completo", anchor=W)
        self.tree.heading("Fornecedor", text="Fornecedor", anchor=W)
        self.tree.heading("Placa", text="Placa", anchor=W)
        self.tree.heading("Local", text="Local", anchor=W)
        self.tree.heading("Nota Fiscal", text="Nota Fiscal", anchor=W)
        self.tree.heading("Hora Chegada", text="Hora Chegada", anchor=W)
        self.tree.heading("Hora Inicial", text="Hora Inicial", anchor=W)
        self.tree.heading("Hora Final", text="Hora Final", anchor=W)
        self.tree.heading("Diferença", text="Diferença", anchor=W)
        self.tree.heading("Observações", text="Observações", anchor=W)
        self.tree.heading("Peso Inicial", text="Peso Inicial", anchor=W)
        self.tree.heading("Tara", text="Tara", anchor=W)
        self.tree.heading("Peso Final", text="Peso Final", anchor=W)
        
        # Carregando dados na Treeview
        self.load_data()
        
        ficheiro = pathlib.Path("Entradas.xlsx")
            
        if ficheiro.exists():
            pass
        else:
            ficheiro=Workbook()
            folha = ficheiro.active
            folha['A1'] = "Nome Completo"
            folha['B1'] = "Fornecedor"
            folha['C1'] = "Placa"
            folha['D1'] = "Local de descarga"
            folha['E1'] = "Nota fiscal"
            folha['F1'] = "Hora Chegada"
            folha['G1'] = "Hora Inicial"
            folha['H1'] = "Hora final"
            folha['I1'] = "Diferença"
            folha['J1'] = "Observações"
            folha['K1'] = "Peso Inicial"
            folha['L1'] = "Tara"
            folha['M1'] = "Peso Final"
                
            ficheiro.save("Entradas.xlsx")
            
        def submit():
            name = name_value.get()
            placa = placa_value.get()
            nf = nf_value.get()
            local = local_combobox.get()
            fornecedor = fornecedor_combobox.get()
            horachegada = horachegada_value.get()
            horainicial = horainicial_value.get()
            horafinal = horafinal_value.get()
            obs = obs_entry.get("1.0", END)  # Ajustado para pegar tudo desde o início até o final
            peso_inicial = peso_inicial_value.get() # Pegando o valor do peso inicial
            tara = tara_value.get() # Pegando o valor da tara
            peso_final = peso_final_value.get() # Pegando o valor do peso final
        
            if (name =="" or fornecedor=="" or placa =="" or nf==""):
                messagebox.showerror("Sistema", "Erro\nPor favor preencha todos os campos.")
            else:
                try:
                    # Calculando a diferença de horas
                    diff = datetime.strptime(horafinal, "%H:%M") - datetime.strptime(horachegada, "%H:%M")
                    diff_hours_minutes = "{:02}:{:02}".format(*divmod(diff.seconds // 60, 60))
                    
                    # Salvar dados no arquivo
                    ficheiro = openpyxl.load_workbook('Entradas.xlsx')
                    folha = ficheiro.active
                    folha.append([name, fornecedor, placa, local, nf, horachegada, horainicial, horafinal, diff_hours_minutes, obs, peso_inicial, tara, peso_final])
                    ficheiro.save("Entradas.xlsx")
                    
                    # Atualizar a Treeview
                    self.tree.delete(*self.tree.get_children())  # Limpa os dados existentes na Treeview
                    self.load_data()  # Carrega novamente os dados na Treeview
                    
                    messagebox.showinfo("Sistema", "Dados salvos com sucesso!")
                except ValueError:
                    messagebox.showerror("Sistema", "Erro\nPor favor, insira as horas no formato correto (HH:MM).")
                
        def alterar():
            item = self.tree.selection()[0]
            name = name_value.get()
            placa = placa_value.get()
            nf = nf_value.get()
            local = local_combobox.get()
            fornecedor = fornecedor_combobox.get()
            horachegada = horachegada_value.get()
            horainicial = horainicial_value.get()
            horafinal = horafinal_value.get()
            obs = obs_entry.get("1.0", END)  # Ajustado para pegar tudo desde o início até o final
            peso_inicial = peso_inicial_value.get() # Pegando o valor do peso inicial
            tara = tara_value.get() # Pegando o valor da tara
            peso_final = peso_final_value.get() # Pegando o valor do peso final
            
            if (name =="" or fornecedor=="" or placa =="" or nf==""):
                messagebox.showerror("Sistema", "Erro\nPor favor preencha todos os campos.")
            else:
                try:
                    # Calculando a diferença de horas
                    diff = datetime.strptime(horafinal, "%H:%M") - datetime.strptime(horachegada, "%H:%M")
                    diff_hours_minutes = "{:02}:{:02}".format(*divmod(diff.seconds // 60, 60))
                    
                    # Atualizar os dados na Treeview
                    self.tree.item(item, values=(name, fornecedor, placa, local, nf, horachegada, horainicial, horafinal, diff_hours_minutes, obs, peso_inicial, tara, peso_final))
                    
                    # Salvar dados no arquivo
                    ficheiro = openpyxl.load_workbook('Entradas.xlsx')
                    folha = ficheiro.active
                    for i, value in enumerate([name, fornecedor, placa, local, nf, horachegada, horainicial, horafinal, diff_hours_minutes, obs, peso_inicial, tara, peso_final], start=1):
                        folha.cell(row=self.tree.index(item) + 2, column=i, value=value)
                    ficheiro.save("Entradas.xlsx")
                    
                    messagebox.showinfo("Sistema", "Dados alterados com sucesso!")
                except ValueError:
                    messagebox.showerror("Sistema", "Erro\nPor favor, insira as horas no formato correto (HH:MM).")
                
        def delete():
            selected_item = self.tree.selection()
            if selected_item:
                for item in selected_item:
                    try:
                        workbook = load_workbook("Entradas.xlsx")
                        sheet = workbook.active
                        
                        # Obtém os valores da linha selecionada na Treeview
                        selected_values = self.tree.item(item, "values")
                        
                        # Procura a linha correspondente na planilha Excel e a exclui
                        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
                            if row == selected_values:
                                index = list(sheet.iter_rows(values_only=True)).index(row)
                                sheet.delete_rows(index + 1)  # Adiciona 2 para corresponder ao índice baseado em 1 do Excel
                                break
                        
                        workbook.save("Entradas.xlsx")
                        messagebox.showinfo("Sucesso", "Item excluído com sucesso.")
                        self.tree.delete(item)  # Remove o item selecionado da Treeview após excluir da planilha
                    except IndexError:
                        messagebox.showerror("Erro", "Por favor, selecione um item para excluir.")
            else:
                messagebox.showwarning("Atenção", "Por favor, selecione um item para excluir.")

        def clear():
            name_value.set("")
            placa_value.set("")
            nf_value.set("")
            horachegada_value.set("")
            horainicial_value.set("")
            horafinal_value.set("")
            obs_entry.delete("1.0", END)
            peso_inicial_value.set("")
            tara_value.set("")
            peso_final_value.set("")
            
        def on_double_click(event):
            item = self.tree.selection()[0]
            values = self.tree.item(item, "values")
            name_value.set(values[0])
            placa_value.set(values[2])
            nf_value.set(values[4])
            local_combobox.set(values[3])
            fornecedor_combobox.set(values[1])
            horachegada_value.set(values[5])  # Defina o valor de horachegada_value corretamente
            horainicial_value.set(values[6])
            horafinal_value.set(values[7])
            obs_entry.delete("1.0", END)
            obs_entry.insert(END, str(values[9]))
            peso_inicial_value.set(values[10])
            tara_value.set(values[11])
            peso_final_value.set(values[12])
        
        self.tree.bind("<Double-1>", on_double_click)  # Ligando o evento de duplo clique à Treeview
        
        #Texts variables
        name_value = StringVar()
        placa_value = StringVar()
        nf_value = StringVar()
        horachegada_value = StringVar()
        horainicial_value = StringVar()
        horafinal_value = StringVar()
        peso_inicial_value = StringVar() # Variável para o peso inicial
        tara_value = StringVar() # Variável para a tara
        peso_final_value = StringVar() # Variável para o peso final
        
        #Entrys
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 16), fg_color="transparent")
        placa_entry = ctk.CTkEntry(self, width=150, textvariable=placa_value, font=("Century Gothic", 16), fg_color="transparent")
        nf_entry = ctk.CTkEntry(self, width=200, textvariable=nf_value, font=("Century Gothic", 16), fg_color="transparent")
        
        # Entrys para campos de hora inicial e final
        horachegada_entry = Entry(self, width=10, font=("Century Gothic", 16), textvariable=horachegada_value)
        horainicial_entry = Entry(self, width=10, font=("Century Gothic", 16), textvariable=horainicial_value)
        horafinal_entry = Entry(self, width=10, font=("Century Gothic", 16), textvariable=horafinal_value)
        
        #Combobox dos fornecedores
        fornecedores = self.load_fornecedores()
        if fornecedores:
            fornecedor_combobox = ctk.CTkComboBox(self, values=fornecedores, font=("Century Gothic bold", 14))
            fornecedor_combobox.set("Selecione o fornecedor")
            fornecedor_combobox.place(x=50, y=220)
        else:
            messagebox.showerror("Erro", "Lista de fornecedores vazia.")
        
        #Combobox
        local_combobox = ctk.CTkComboBox(self, values=["Produção", "Logística"], font=("Century Gothic bold", 14))
        local_combobox.set("Produção")
        
        #Entrada de observações
        obs_entry = ctk.CTkTextbox(self, width=450, height=60, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent")
        
        # Entrys para peso inicial, tara e peso final
        peso_inicial_entry = ctk.CTkEntry(self, width=50, textvariable=peso_inicial_value, font=("Century Gothic", 16), fg_color="transparent")
        tara_entry = ctk.CTkEntry(self, width=50, textvariable=tara_value, font=("Century Gothic", 16), fg_color="transparent")
        peso_final_entry = ctk.CTkEntry(self, width=50, textvariable=peso_final_value, font=("Century Gothic", 16), fg_color="transparent")
        
        #Labels
        lb_name = ctk.CTkLabel(self, text="Nome completo: ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_placa = ctk.CTkLabel(self, text="Placa da carreta: ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_local = ctk.CTkLabel(self, text="Local ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_fornecedor = ctk.CTkLabel(self, text="Fornecedor ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_nf = ctk.CTkLabel(self, text="Nota fiscal ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_horachegada = ctk.CTkLabel(self, text="Hora Chegada ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_horainicial = ctk.CTkLabel(self, text="Hora Inicial ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_horafinal = ctk.CTkLabel(self, text="Hora Final ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observação ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        lb_peso_inicial = ctk.CTkLabel(self, text="Peso Inicial: ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_tara = ctk.CTkLabel(self, text="Tara: ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_peso_final = ctk.CTkLabel(self, text="Peso Final: ", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        
        btn_submit = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131")
        btn_submit.place(x=200, y=420)
        
        btn_alterar = ctk.CTkButton(self, text="Alterar dados".upper(), command=alterar, fg_color="#151", hover_color="#131")
        btn_alterar.place(x=400, y=420)
        
        btn_delete = ctk.CTkButton(self, text="Excluir item".upper(), command=delete, fg_color="#f00", hover_color="#800")
        btn_delete.place(x=600, y=420)
        
        btn_clear = ctk.CTkButton(self, text="Limpar campos".upper(), command=clear, fg_color="#555", hover_color="#333")
        btn_clear.place(x=800, y=420)
        
        btn_visualizar = ctk.CTkButton(self, text="Visualizar Dashboard".upper(), command=self.visualizar_dashboard, fg_color="#00f", hover_color="#005")
        btn_visualizar.place(x=1000, y=420)
        
        # Posicionando os elementos na janela
        lb_name.place(x=50, y=110)
        name_entry.place(x=50, y=140)
        
        lb_placa.place(x=450, y=110)
        placa_entry.place(x=450, y=140)
        
        lb_local.place(x=250, y=190)
        local_combobox.place(x=250, y=220)
        
        lb_fornecedor.place(x=50, y=190)
        
        lb_nf.place(x=450, y=190)
        nf_entry.place(x=450, y=220)
        
        lb_obs.place(x=50, y=310)
        obs_entry.place(x=50, y=340)
        
        lb_horachegada.place(x=50, y=260)
        horachegada_entry.place(x=160, y=260)
        
        lb_horainicial.place(x=290, y=260)
        horainicial_entry.place(x=370, y=260)
        
        lb_horafinal.place(x=500, y=260)
        horafinal_entry.place(x=580, y=260)
        
        lb_peso_inicial.place(x=750, y=110)
        peso_inicial_entry.place(x=850, y=110)
        
        lb_tara.place(x=750, y=160)
        tara_entry.place(x=800, y=160)
        
        lb_peso_final.place(x=750, y=210)
        peso_final_entry.place(x=850, y=210)
        
    def visualizar_dashboard(self):
        subprocess.Popen(["python", "dashboard.py"])
        
    def change_appearance(self, appearance):
        ctk.set_default_color_theme(appearance.lower())
        self.update_theme()

app = App()
app.mainloop()
